"""S9a — NHS England Discharge Ready Date (DRD) monthly ETL.

Reconstructed 2026-08-14. The original build ran ad hoc in July 2026 and its
code was never committed: 3,978 rows were live, wired into Workflow 1 and
driving the mental_health and learning_disability tenant types, with no code
path back to them. Node documentation described what was done but is not a
specification and is not executable.

This script is rebuilt from docs/nodes/s9a_node1..s9a_node3 and is verified by
exact reproduction: --reproduce loads into a staging table and diffs every
cell against the live table. Anything that does not reproduce is either a
reconstruction error or something the original transform did and did not
write down, and the difference has to be resolved before either is trusted.

Reproduction is possible at all because nhs_drd_discharge_delays records the
source URL on every row, so the rebuild can be pointed at exactly the files
that produced the live data rather than at whatever the page serves today.

Usage:
    python scripts/s9a_drd_build.py --reproduce     # staging + exact diff
    python scripts/s9a_drd_build.py --load          # upsert the live table
    python scripts/s9a_drd_build.py --load --from 2026-06-01
"""
import argparse
import re
import sys
import urllib.request
from pathlib import Path

import openpyxl
import psycopg2.extras

sys.path.insert(0, str(Path(__file__).resolve().parent))
from _db import get_conn  # noqa: E402

if hasattr(sys.stdout, "reconfigure"):
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")

REPO = Path(__file__).resolve().parent.parent
RAW = REPO / "data" / "raw" / "s9a_drd"

# Corrected 2026-08-14. This was previously documented against the acute
# discharge sitrep page, which is a different publication and stops at
# September 2024.
PAGE = ("https://www.england.nhs.uk/statistics/statistical-work-areas/"
        "discharge-delays/discharge-ready-date/")
LINK_RE = r'href="([^"]*Discharge-Ready-Date[^"]*\.xlsx)"'
UA = {"User-Agent": "ucws-pipeline/s9a (+sl@slendeavours.org)"}

MONTHS = ("january february march april may june july august september "
          "october november december").split()
MONTH_RE = "|".join(MONTHS)

TABLE = "nhs_drd_discharge_delays"
STAGING = "nhs_drd_discharge_delays_repro"

# Column mapping, 0-indexed, from docs/nodes/s9a_node2_parse_utla_aggregate.md.
# Column 9 (total_discharges) is always null at UTLA Aggregate level — it is a
# per-provider field — and is carried anyway so the shape matches the table.
COLS = [
    ("utla_code", 1), ("utla_name", 2),
    ("pct_acceptable_trust_coverage", 8),
    ("total_discharges", 9),
    ("total_discharges_acceptable_trusts", 10),
    ("total_bed_days_lost", 11),
    ("pct_same_day_discharge", 13),
    ("pct_delayed_1plus_days", 14),
    ("discharged_no_delay", 16), ("discharged_1_day", 17),
    ("discharged_2_3_days", 18), ("discharged_4_6_days", 19),
    ("discharged_7_13_days", 20), ("discharged_14_20_days", 21),
    ("discharged_21_plus_days", 22),
    ("avg_days_drd_to_discharge_inc_zero", 46),
    ("avg_days_drd_to_discharge_exc_zero", 47),
]
FIELDS = [c for c, _ in COLS]
HEADER_ROW_0IDX = 14          # header at row 15, data from row 16
SUMMARY_TYPE_COL = 0
SUPPRESSION = {"-", "*", "..", ""}


def halt(msg):
    sys.exit(f"HALT: {msg}")


def parse_period(url):
    m = re.search(rf"({MONTH_RE})[_\-]?(\d{{4}})", url.lower())
    if not m:
        return None
    return f"{int(m.group(2)):04d}-{MONTHS.index(m.group(1)) + 1:02d}-01"


def discover():
    """period -> file URL, newest wins where a month appears twice."""
    req = urllib.request.Request(PAGE, headers=UA)
    with urllib.request.urlopen(req, timeout=60) as r:
        body = r.read().decode("utf-8", errors="replace")
    found = {}
    for link in re.findall(LINK_RE, body, re.I):
        if "timeseries" in link.lower():
            continue          # the timeseries workbook is not a monthly file
        period = parse_period(link)
        if not period:
            continue
        # A -Revised file supersedes the original for the same month.
        if period not in found or "revis" in link.lower():
            found[period] = link
    if not found:
        halt("no monthly DRD links matched on the publication page")
    return found


def download(url):
    RAW.mkdir(parents=True, exist_ok=True)
    dest = RAW / url.rsplit("/", 1)[-1]
    if dest.exists() and dest.stat().st_size > 10_000:
        return dest
    req = urllib.request.Request(url, headers=UA)
    with urllib.request.urlopen(req, timeout=180) as r, open(dest, "wb") as f:
        f.write(r.read())
    if dest.stat().st_size < 10_000:
        halt(f"{dest.name} downloaded under 10 KB")
    return dest


def coerce(value):
    """Suppression markers become NULL, never zero."""
    if value is None:
        return None
    if isinstance(value, str):
        v = value.strip()
        if v in SUPPRESSION:
            return None
        try:
            return float(v.replace(",", "").replace("%", ""))
        except ValueError:
            return v
    return value


def parse_file(path, period, url):
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    sheet = next((s for s in wb.sheetnames
                  if re.search(r"utla.*acceptable", s, re.I)), None)
    if not sheet:
        halt(f"{path.name}: no sheet matching *UTLA*Acceptable* "
             f"(sheets: {wb.sheetnames})")
    ws = wb[sheet]

    rows = []
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i <= HEADER_ROW_0IDX:
            continue
        if not row or len(row) <= max(c for _, c in COLS):
            continue
        summary = (row[SUMMARY_TYPE_COL] or "")
        if str(summary).strip() != "UTLA Aggregate":
            continue
        code = row[1]
        if not code or not str(code).strip().startswith("E"):
            continue
        rec = {"reporting_period": period, "source": url}
        for name, idx in COLS:
            v = coerce(row[idx])
            if name in ("utla_code", "utla_name"):
                rec[name] = str(v).strip() if v is not None else None
            else:
                rec[name] = v
        rows.append(rec)
    wb.close()
    return rows


DDL_STAGING = f"""
CREATE TABLE IF NOT EXISTS {STAGING} (LIKE {TABLE} INCLUDING ALL);
TRUNCATE {STAGING};
"""

INSERT_COLS = (["reporting_period"] + FIELDS + ["source"])


def write_rows(cur, table, rows):
    placeholders = ", ".join(["%s"] * len(INSERT_COLS))
    updates = ", ".join(f"{c} = EXCLUDED.{c}" for c in INSERT_COLS
                        if c not in ("reporting_period", "utla_code"))
    sql = (f"INSERT INTO {table} ({', '.join(INSERT_COLS)}) "
           f"VALUES ({placeholders}) "
           f"ON CONFLICT (reporting_period, utla_code) DO UPDATE SET "
           f"{updates}, loaded_at = now()")
    psycopg2.extras.execute_batch(
        cur, sql, [tuple(r[c] for c in INSERT_COLS) for r in rows])


NUMERIC = [f for f in FIELDS if f not in ("utla_code", "utla_name")]


def diff(cur):
    """Cell-level diff of the staging rebuild against the live table."""
    cur.execute(f"SELECT COUNT(*) FROM {TABLE}")
    live_n = cur.fetchone()[0]
    cur.execute(f"SELECT COUNT(*) FROM {STAGING}")
    repro_n = cur.fetchone()[0]

    cur.execute(f"""
        SELECT COUNT(*) FROM (
            SELECT reporting_period, utla_code FROM {TABLE}
            EXCEPT SELECT reporting_period, utla_code FROM {STAGING}) d
    """)
    only_live = cur.fetchone()[0]
    cur.execute(f"""
        SELECT COUNT(*) FROM (
            SELECT reporting_period, utla_code FROM {STAGING}
            EXCEPT SELECT reporting_period, utla_code FROM {TABLE}) d
    """)
    only_repro = cur.fetchone()[0]

    mismatches = []
    for col in ["utla_name"] + NUMERIC:
        # IS DISTINCT FROM so NULL vs NULL counts as equal and NULL vs 0 does
        # not — suppression handling is exactly what has to reproduce.
        cur.execute(f"""
            SELECT COUNT(*) FROM {TABLE} a JOIN {STAGING} b
              ON a.reporting_period = b.reporting_period
             AND a.utla_code = b.utla_code
            WHERE a."{col}" IS DISTINCT FROM b."{col}"
        """)
        n = cur.fetchone()[0]
        if n:
            cur.execute(f"""
                SELECT a.reporting_period, a.utla_code, a."{col}", b."{col}"
                FROM {TABLE} a JOIN {STAGING} b
                  ON a.reporting_period = b.reporting_period
                 AND a.utla_code = b.utla_code
                WHERE a."{col}" IS DISTINCT FROM b."{col}" LIMIT 3
            """)
            mismatches.append((col, n, cur.fetchall()))
    return live_n, repro_n, only_live, only_repro, mismatches


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--reproduce", action="store_true",
                    help="rebuild into a staging table and diff it")
    ap.add_argument("--load", action="store_true", help="upsert the live table")
    ap.add_argument("--from", dest="since", help="earliest period, YYYY-MM-DD")
    args = ap.parse_args()
    if not (args.reproduce or args.load):
        halt("choose --reproduce or --load")

    conn = get_conn()
    conn.autocommit = False
    cur = conn.cursor()

    published = discover()
    print(f"discovered {len(published)} monthly file(s) on the publication page")

    if args.reproduce:
        # Reproduce against the files that produced the live rows, taken from
        # the table's own source column, not from whatever the page serves now.
        cur.execute(f"SELECT DISTINCT reporting_period::text, source "
                    f"FROM {TABLE} ORDER BY 1")
        targets = dict(cur.fetchall())
        print(f"reproducing {len(targets)} period(s) from recorded source URLs")
        drift = [p for p, u in targets.items()
                 if published.get(p) and published[p] != u]
        if drift:
            print(f"  note: {len(drift)} period(s) now serve a different file "
                  f"than was loaded: {drift[:4]}")
    else:
        targets = {p: u for p, u in published.items()
                   if not args.since or p >= args.since}
        print(f"loading {len(targets)} period(s)")

    all_rows = []
    for period in sorted(targets):
        url = targets[period]
        path = download(url)
        rows = parse_file(path, period, url)
        all_rows.extend(rows)
        print(f"  {period}  {len(rows):>4} UTLA rows  {path.name[:58]}")

    if not all_rows:
        halt("no rows parsed")

    try:
        if args.reproduce:
            cur.execute(DDL_STAGING)
            write_rows(cur, STAGING, all_rows)
            live_n, repro_n, only_live, only_repro, mismatches = diff(cur)
            print()
            print(f"live rows      : {live_n}")
            print(f"rebuilt rows   : {repro_n}")
            print(f"keys only live : {only_live}")
            print(f"keys only repro: {only_repro}")
            if mismatches:
                print(f"columns differing: {len(mismatches)}")
                for col, n, sample in mismatches:
                    print(f"  {col}: {n} row(s)")
                    for s in sample:
                        print(f"     {s[0]} {s[1]}: live={s[2]!r} repro={s[3]!r}")
            else:
                print("columns differing: 0")
            ok = (live_n == repro_n and not only_live and not only_repro
                  and not mismatches)
            conn.commit()
            print()
            print("EXACT REPRODUCTION" if ok else "REPRODUCTION FAILED")
            return 0 if ok else 1
        else:
            write_rows(cur, TABLE, all_rows)
            cur.execute(f"SELECT COUNT(*) FROM {TABLE}")
            print(f"{TABLE}: {cur.fetchone()[0]} rows")
            conn.commit()
            print("COMMITTED")
            return 0
    except Exception:
        conn.rollback()
        raise
    finally:
        cur.close()
        conn.close()


if __name__ == "__main__":
    sys.exit(main())
