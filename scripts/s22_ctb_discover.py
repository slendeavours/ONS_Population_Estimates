"""S22 Phase 1 — discover the two MHCLG source files from their landing pages.

No file URL is hardcoded. Both publishers are resolved at runtime through the
GOV.UK content API, which is the machine-readable form of the same landing
pages a human would read:

  Source A  Council Taxbase statistics collection
            -> most recent "Council Taxbase <year> in England" release
            -> the local-authority-level data workbook attached to it
  Source B  Live tables on dwelling stock including vacants
            -> Table 615, vacant dwellings by local authority district

Writes build_reports/s22_source_structure.md and returns the resolved paths.
"""
import datetime
import json
import re
import sys
import urllib.request
from pathlib import Path

import openpyxl
import pandas as pd

REPO = Path(__file__).resolve().parent.parent
RAW_DIR = REPO / "data" / "raw" / "s22_ctb"
REPORT_DIR = REPO / "build_reports"

API = "https://www.gov.uk/api/content"
COLLECTION_PATH = "/government/collections/council-taxbase-statistics"
LIVE_TABLES_PATH = ("/government/statistical-data-sets/"
                    "live-tables-on-dwelling-stock-including-vacants")

UA = {"User-Agent": "ucws-pipeline/s22 (+sl@slendeavours.org)"}


def halt(msg):
    sys.exit(f"HALT: {msg}")


def _get_json(path):
    req = urllib.request.Request(API + path, headers=UA)
    with urllib.request.urlopen(req, timeout=120) as r:
        return json.loads(r.read().decode("utf-8"))


def _download(url, dest):
    if dest.exists() and dest.stat().st_size > 0:
        return dest
    dest.parent.mkdir(parents=True, exist_ok=True)
    req = urllib.request.Request(url, headers=UA)
    with urllib.request.urlopen(req, timeout=600) as r, open(dest, "wb") as f:
        f.write(r.read())
    return dest


# ── Source A: Council Taxbase ───────────────────────────────────────────────

def discover_council_taxbase():
    coll = _get_json(COLLECTION_PATH)
    links = []
    for group in coll.get("links", {}).get("documents", []):
        title = group.get("title", "")
        m = re.match(r"Council Taxbase (\d{4}) in England\s*$", title)
        if m:
            links.append((int(m.group(1)), title, group["base_path"],
                          group.get("public_updated_at")))
    if not links:
        halt("no 'Council Taxbase <year> in England' release found on "
             f"{COLLECTION_PATH} — landing page structure has changed")
    links.sort(reverse=True)
    year, title, base_path, _ = links[0]

    rel = _get_json(base_path)
    attachments = rel["details"].get("attachments", [])
    data_file = None
    for a in attachments:
        t = a.get("title", "")
        if re.search(r"local authority level data", t, re.I) and a.get("url"):
            data_file = a
            break
    if data_file is None:
        halt(f"no local-authority-level data file attached to '{title}' — "
             f"attachments seen: {[a.get('title') for a in attachments]}")

    tech = next((a for a in attachments
                 if re.search(r"technical note", a.get("title", ""), re.I)),
                None)
    tech_url = tech.get("url") if tech else None
    if tech_url and tech_url.startswith("/"):
        tech_url = "https://www.gov.uk" + tech_url

    dest = RAW_DIR / Path(data_file["url"]).name
    _download(data_file["url"], dest)

    return {
        "release_title": rel["title"],
        "release_page": "https://www.gov.uk" + base_path,
        "taxbase_year": year,
        "first_published": rel.get("first_published_at"),
        "public_updated": rel.get("public_updated_at"),
        "attachment_title": data_file["title"],
        "url": data_file["url"],
        "content_type": data_file.get("content_type"),
        "file_size": data_file.get("file_size"),
        "path": dest,
        "technical_notes_url": tech_url,
        "all_attachments": [a.get("title") for a in attachments],
    }


# ── Source B: Table 615 ─────────────────────────────────────────────────────

def discover_table_615():
    page = _get_json(LIVE_TABLES_PATH)
    att = None
    for a in page["details"].get("attachments", []):
        if re.match(r"Table 615\b", a.get("title", "")) and a.get("url"):
            att = a
            break
    if att is None:
        halt("Table 615 not found on the live tables on dwelling stock "
             "landing page — page structure has changed")
    dest = RAW_DIR / Path(att["url"]).name
    _download(att["url"], dest)
    return {
        "release_title": page["title"],
        "release_page": "https://www.gov.uk" + LIVE_TABLES_PATH,
        "attachment_title": att["title"],
        "url": att["url"],
        "content_type": att.get("content_type"),
        "file_size": att.get("file_size"),
        "public_updated": page.get("public_updated_at"),
        "path": dest,
    }


# ── Structure report ────────────────────────────────────────────────────────

def _ctb_sheet_structure(path):
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    out = []
    for name in wb.sheetnames:
        ws = wb[name]
        out.append({"sheet": name, "rows": ws.max_row, "cols": ws.max_column})
    wb.close()
    return out


def _615_sheet_structure(path):
    out = []
    for sheet in ("All_vacants", "All_long_term_vacants"):
        df = pd.read_excel(path, sheet_name=sheet, engine="odf", header=None)
        out.append({"sheet": sheet, "rows": df.shape[0], "cols": df.shape[1],
                    "title": str(df.iloc[0, 0]),
                    "header_row": 3,
                    "headers": [str(v) for v in df.iloc[2].tolist()]})
    return out


def write_structure_report(src_a, src_b, ctb_tables, national, sheets_a,
                           sheets_b, exemption_classes, source_number):
    REPORT_DIR.mkdir(parents=True, exist_ok=True)
    p = REPORT_DIR / f"s{source_number}_source_structure.md"
    now = datetime.datetime.now(datetime.timezone.utc).isoformat(
        timespec="seconds")

    L = []
    a = L.append
    a(f"# S{source_number} — source structure report")
    a("")
    a(f"Generated {now}. Both files resolved at runtime from their publisher "
      "landing pages via the GOV.UK content API; no file URL is hardcoded in "
      "this build.")
    a("")

    a("## Source A — Local authority Council Taxbase in England")
    a("")
    a(f"| field | value |")
    a("|---|---|")
    a(f"| landing page | {src_a['release_page']} |")
    a(f"| release title | {src_a['release_title']} |")
    a(f"| taxbase year | {src_a['taxbase_year']} |")
    a(f"| first published | {src_a['first_published']} |")
    a(f"| revised / last updated | {src_a['public_updated']} |")
    a(f"| attachment | {src_a['attachment_title']} |")
    a(f"| resolved URL | {src_a['url']} |")
    a(f"| format | {src_a['content_type']} |")
    a(f"| file size | {src_a['file_size']:,} bytes |")
    a(f"| technical notes | {src_a['technical_notes_url']} |")
    a("")
    a("Attachments on the release page:")
    a("")
    for t in src_a["all_attachments"]:
        a(f"- {t}")
    a("")
    a("### Sheets")
    a("")
    a("| sheet | rows | columns |")
    a("|---|---|---|")
    for s in sheets_a:
        a(f"| {s['sheet']} | {s['rows']} | {s['cols']} |")
    a("")
    a("Header row position: **row 6** on every data sheet. Row 5 carries the "
      "table label spanning each block, row 6 the column headers, row 7 the "
      "England total, rows 8-303 the 296 billing authorities.")
    a("")
    a("Identity columns (row 6, 0-indexed): 0 `E-code`, 1 `ONS Code`, "
      "2 `Region`, 3 `Local Authority`, 4 `Notes`.")
    a("")
    a("### Tables used, as published")
    a("")
    a("| table | column header used | 0-indexed column | published title |")
    a("|---|---|---|---|")
    for t in ctb_tables:
        a(f"| {t['table']} | {t['header']} | {t['col']} | {t['title']} |")
    a("")
    a("Table 2.01 on the `Supplementary Data` sheet publishes the exemption "
      "class breakdown **at local authority level**, one column per class "
      "A to W plus `Total exemptions`. Classes used for this build: "
      + ", ".join(exemption_classes) + ".")
    a("")
    a("Only the current taxbase year is present in this workbook. No "
      "prior-year columns are published in the same file, so a single year "
      f"({src_a['taxbase_year']}) is loaded.")
    a("")

    a("## Source B — Table 615, vacant dwellings by local authority district")
    a("")
    a("| field | value |")
    a("|---|---|")
    a(f"| landing page | {src_b['release_page']} |")
    a(f"| attachment | {src_b['attachment_title']} |")
    a(f"| resolved URL | {src_b['url']} |")
    a(f"| format | {src_b['content_type']} |")
    a(f"| file size | {src_b['file_size']:,} bytes |")
    a(f"| landing page last updated | {src_b['public_updated']} |")
    a("")
    for s in sheets_b:
        a(f"### Sheet `{s['sheet']}`")
        a("")
        a(f"{s['title']}")
        a("")
        a(f"- {s['rows']} rows x {s['cols']} columns")
        a(f"- header row position: row {s['header_row']}")
        a("- column headers as published: `"
          + "`, `".join(s["headers"]) + "`")
        a("")
    a("Each data column header is the snapshot date for that year. Rows "
      "include England (`E92000001`) and the nine regions (`E12...`) as well "
      "as districts; only `E06`/`E07`/`E08`/`E09` rows are loaded. Suppressed "
      "or not-applicable cells are published as `[x]`.")
    a("")

    a("## National headline figures printed on the release page")
    a("")
    a("These are the Phase 6 reconciliation targets. They are taken from the "
      "MHCLG statistical release text itself, not from any secondary "
      "commentary.")
    a("")
    a("| figure | as printed on the release page | value |")
    a("|---|---|---|")
    for row in national:
        a(f"| {row['figure']} | {row['quote']} | {row['value']} |")
    a("")
    a("**Not printed on the release page:** the release states no national "
      "figure for dwellings empty for more than six months, and none for "
      "dwellings empty for less than six months. Those two are NOT FOUND on "
      "the release page — they are not unchecked. For those two only, the "
      "reconciliation target is the publisher's own England total row (row 7) "
      "in the same local-authority-level workbook, which is stated in the "
      "verification report wherever the figure appears.")
    a("")

    a("## Structural breaks stated by the publisher")
    a("")
    a("Both are stated in the release text at "
      f"{src_a['release_page']} and in the technical notes at "
      f"{src_a['technical_notes_url']}:")
    a("")
    a("- **1 April 2024** — \"authorities could charge an Empty Homes Premium "
      "of up to 100% for properties that have been empty for between 1 and 2 "
      "years. Previously the premium could only be applied where properties "
      "had been empty for 2 or more years.\" Premium counts are not "
      "comparable across this date.")
    a("- **1 April 2025** — \"authorities could charge a Second Homes Premium "
      "of up to 100% on properties that were reported to be second homes for "
      "council tax purposes.\" Second home counts from this date are affected "
      "by reclassification behaviour.")
    a("")
    p.write_text("\n".join(L), encoding="utf-8")
    return p
