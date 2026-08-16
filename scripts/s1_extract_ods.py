"""S1 — extract the statutory homelessness ODS to the load-ready CSV.

This step had no committed code anywhere. S1's n8n node 1 fetches a
pre-processed CSV from the public repository; the ODS-to-CSV conversion that
produced that CSV was done locally and never written down. That is why 2025Q1
could not be reloaded: the loader works, and it fetches a CSV that does not
exist for that quarter.

Streaming, not DOM. content.xml is ~42 MB uncompressed for a single quarter
and the workbook carries 51 sheets, so it is parsed with iterparse over the
zip entry, honouring number-columns-repeated and number-rows-repeated, and
cleared as it goes. Same approach as scripts/s11_cqc_fetch.py, which hit the
same wall on the CQC directory.

Only three sheets are read: A1 (assessments and duties), TA1 (temporary
accommodation) and A3 (support needs). The rest are skipped without parsing
their rows.

Usage:
    python scripts/s1_extract_ods.py --period 2025Q3
    python scripts/s1_extract_ods.py --period 2025Q1 --out data/processed/x.csv
"""
import argparse
import csv
import re
import sys
import xml.etree.ElementTree as ET
import zipfile
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent))

if hasattr(sys.stdout, "reconfigure"):
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")

REPO = Path(__file__).resolve().parent.parent
RAW_DIR = REPO / "data" / "raw" / "s1b_a3"

TABLE = "{urn:oasis:names:tc:opendocument:xmlns:table:1.0}"
OFFICE = "{urn:oasis:names:tc:opendocument:xmlns:office:1.0}"
TEXT = "{urn:oasis:names:tc:opendocument:xmlns:text:1.0}"

WANTED_SHEETS = {"A1", "TA1", "A3"}


def halt(msg):
    sys.exit(f"HALT: {msg}")


def cell_value(cell):
    """Prefer the typed value; fall back to displayed text.

    office:value carries the number as stored. The text runs carry what is
    displayed, which for a suppressed or not-applicable cell is a marker such
    as '..' or '-' rather than a number. Reading the typed value first and the
    text second keeps those markers visible instead of coercing them to zero.
    """
    v = cell.get(OFFICE + "value")
    if v is not None:
        return v
    parts = []
    for p in cell.iter(TEXT + "p"):
        parts.append("".join(p.itertext()))
    return " ".join(x.strip() for x in parts if x is not None).strip()


def read_sheets_ods(ods_path, wanted):
    """{sheet name: [row, ...]} for the wanted sheets only."""
    out = {}
    z = zipfile.ZipFile(ods_path)
    with z.open("content.xml") as f:
        sheet = None
        rows = None
        for event, elem in ET.iterparse(f, events=("start", "end")):
            if event == "start" and elem.tag == TABLE + "table":
                sheet = elem.get(TABLE + "name")
                rows = [] if sheet in wanted else None
            elif event == "end" and elem.tag == TABLE + "table-row":
                if rows is None:
                    elem.clear()
                    continue
                row = []
                for cell in elem:
                    if cell.tag not in (TABLE + "table-cell",
                                        TABLE + "covered-table-cell"):
                        continue
                    rep = int(cell.get(TABLE + "number-columns-repeated", "1"))
                    val = cell_value(cell)
                    if rep > 1000 and val == "":
                        rep = 1          # trailing filler columns
                    row.extend([val] * rep)
                while row and row[-1] == "":
                    row.pop()
                if row:
                    rrep = int(elem.get(TABLE + "number-rows-repeated", "1"))
                    for _ in range(min(rrep, 1000)):
                        rows.append(list(row))
                elem.clear()
            elif event == "end" and elem.tag == TABLE + "table":
                if rows is not None:
                    out[sheet] = rows
                sheet, rows = None, None
                elem.clear()
                if len(out) == len(wanted):
                    break
    return out


def find_header(rows, must_contain):
    """Index of the header row, located by content rather than a fixed offset.

    MHCLG moves the header between editions - the row index that worked for
    one quarter has not worked for the next - so it is found by looking for
    the row that carries the ONS code column.
    """
    for i, row in enumerate(rows[:40]):
        joined = " | ".join(row).lower()
        if all(t.lower() in joined for t in must_contain):
            return i
    return None


def la_rows(rows, header_idx):
    """Rows whose first cell that looks like a GSS code is an English LA."""
    out = []
    for row in rows[header_idx + 1:]:
        code = next((c.strip() for c in row[:4]
                     if re.fullmatch(r"E0[6-9]\d{6}", (c or "").strip())), None)
        if code:
            out.append((code, row))
    return out


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--period", required=True, help="e.g. 2025Q3")
    ap.add_argument("--ods", help="explicit path, otherwise resolved by period")
    ap.add_argument("--out", help="output CSV path")
    ap.add_argument("--source", help="source label for provenance")
    ap.add_argument("--inspect", action="store_true",
                    help="report sheet shapes and headers, write nothing")
    args = ap.parse_args()

    if args.ods:
        ods = Path(args.ods)
    else:
        matches = sorted(RAW_DIR.glob(f"{args.period}_*"))
        if not matches:
            halt(f"no raw file for {args.period} under {RAW_DIR}")
        if len(matches) > 1:
            halt(f"{len(matches)} raw files match {args.period}: "
                 f"{[m.name for m in matches]}. Pass --ods explicitly.")
        ods = matches[0]
    if not ods.exists():
        halt(f"{ods} not found")
    print(f"period : {args.period}")
    print(f"source : {ods.name}")

    sheets = read_sheets(ods, WANTED_SHEETS)
    missing = WANTED_SHEETS - set(sheets)
    if missing:
        halt(f"sheets not found in the workbook: {sorted(missing)}")

    for name in sorted(sheets):
        rows = sheets[name]
        las = [r for r in rows if r and LA_CODE.fullmatch((r[0] or "").strip())]
        print(f"  {name:<4} {len(rows):>5} rows, {len(las)} English LA rows")

    if args.inspect:
        return 0

    source_label = args.source or f"MHCLG Statutory Homelessness Detailed Local Authority Data, {ods.name}"
    recs = extract(ods, args.period, source_label)
    out = Path(args.out) if args.out else (
        REPO / "data" / "processed" / f"statutory_homelessness_{args.period}.csv")
    out.parent.mkdir(parents=True, exist_ok=True)
    with open(out, "w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=OUT_FIELDS)
        w.writeheader()
        for r in recs:
            w.writerow({k: ("" if r.get(k) is None else r.get(k))
                        for k in OUT_FIELDS})
    try:
        shown = out.relative_to(REPO)
    except ValueError:
        shown = out
    print(f"  wrote {shown}: {len(recs)} LA rows")
    return 0



# Columns are resolved by header text, never by a fixed index. The five
# support-need columns were quarantined precisely because a fixed offset was
# used and the offset varies by quarter: the stored values held different
# publisher columns from the ones their names claimed. Matching the header
# label is the fix, and it fails loudly when a label changes rather than
# silently reading the neighbouring column.
COLUMN_LABELS = {
    # Ordered most-specific first. The 2025Q4 edition renamed and moved every
    # column in all three sheets, so each measure carries the labels for both
    # the pre-2025Q4 layout and the current one.
    "A1": {
        "total_assessments": [
            "initial assessments total",                      # 2025Q4-
            "total initial assessments"],                     # -2025Q3
        "owed_duty": [
            "assessed as owed a duty owed a prevention or relief duty total",
            "total owed a prevention or relief",
            "assessed as owed a duty"],
        "prevention_duty": [
            "assessed as owed a duty threatened with homelessness within 56 days",
            "threatened with homelessness within 56 days"],
        "relief_duty": [
            "assessed as owed a duty homeless, relief duty owed",
            "homeless - relief duty owed",
            "relief duty owed"],
    },
    "TA1": {
        "households_in_ta": [
            "households in ta total",             # 2025Q4-
            "sum of total households in ta",      # -2025Q3, unique; the
            # plain 'total number of households in TA' prefix also matches the
            # per-000s and with-children columns, so it is not used.
            "households in temporary accommodation at end"],
    },
    "A3": {
        "support_needs_total": ["households with one or more support needs"],
        "mental_health": ["history of mental health problems"],
        "learning_disability": ["learning disability"],
        "drug_dependency": ["drug dependency needs"],
        "alcohol_dependency": ["alcohol dependency needs"],
        "rough_sleeping_history": ["history of rough sleeping"],
    },
}

OUT_FIELDS = ["lad24cd", "la_name", "period", "total_assessments", "owed_duty",
              "prevention_duty", "relief_duty", "households_in_ta",
              "support_needs_total", "mental_health", "learning_disability",
              "drug_dependency", "alcohol_dependency",
              "rough_sleeping_history", "source"]

LA_CODE = re.compile(r"E0[6-9]\d{6}")



def code_resolution():
    """Publisher GSS code -> canonical lad24cd, from la_code_lookup.

    MHCLG publishes Barnsley and Sheffield on the codes recoded on 1 April
    2025 (SI 1328/2024), E08000038 and E08000039, while la_boundaries is LAD
    May 2024 and carries E08000016 and E08000019. Resolution is
    change_type = 'recode' only: a recode renumbers the same area, while
    new_unitary and merger are abolitions and must stay unmapped, because
    folding predecessor districts onto a successor makes every downstream sum
    count that successor once per predecessor.
    """
    from _db import get_readonly_conn
    conn = get_readonly_conn()
    cur = conn.cursor()
    cur.execute("SELECT lad24cd FROM la_boundaries")
    current = {r[0] for r in cur.fetchall()}
    cur.execute("""SELECT old_code, new_code FROM la_code_lookup
                   WHERE change_type = 'recode'""")
    mapping = {o: n for o, n in cur.fetchall() if n in current}
    cur.close()
    conn.close()
    return current, mapping



def read_sheets_xlsx(xlsx_path, wanted):
    """Same shape, for the quarters MHCLG published as .xlsx.

    MHCLG changed container format mid-series - 2023Q4, 2024Q1 and 2024Q2 are
    .xlsx while everything either side is .ods. The format is not a property
    of the data, so it must not be a property of the pipeline: both readers
    return the same {sheet: [row, ...]} and everything downstream is shared.

    data_only=True reads cached formula results. A suppression marker is a
    literal string in the cell either way, so markers survive here exactly as
    they do in the ODS path.
    """
    import openpyxl
    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    out = {}
    for name in wb.sheetnames:
        if name not in wanted:
            continue
        rows = []
        for r in wb[name].iter_rows(values_only=True):
            row = ["" if c is None else str(c).strip() for c in r]
            while row and row[-1] == "":
                row.pop()
            if row:
                rows.append(row)
        out[name] = rows
    wb.close()
    return out


def read_sheets(path, wanted):
    """Dispatch on the container, and say so rather than dying in zipfile."""
    suffix = path.suffix.lower()
    if suffix == ".ods":
        return read_sheets_ods(path, wanted)
    if suffix in (".xlsx", ".xlsm"):
        return read_sheets_xlsx(path, wanted)
    halt(f"unsupported source container '{suffix}' for {path.name}")



def is_rate_header(cell):
    """True for per-thousand and thousands columns.

    From 2025Q4 the sheets carry rate columns next to the counts, and
    'Households in TA  Total  (per thousand)  Total' begins with the same text
    as the count column. These tables are loaded as counts, so rate columns
    are never a valid answer and are excluded before matching rather than
    disambiguated afterwards.
    """
    c = " ".join((cell or "").lower().split())
    return ("per thousand" in c or "(thousands)" in c
            or "per (000s)" in c or "(000s)" in c)


def norm_header(cell):
    """Lowercased, whitespace-collapsed header text with the edition's own
    prefix removed.

    From 2025Q4 MHCLG prefixes every A3 support-need column with 'Support
    need', so 'History of mental health problems' became 'Support need
    History of mental health problems'. The prefix is presentation, not
    meaning, and stripping it lets one label list span both editions.
    """
    c = " ".join((cell or "").lower().split())
    for prefix in ("support need ", "of which: "):
        if c.startswith(prefix):
            c = c[len(prefix):]
    return c


def resolve_columns(rows, labels, sheet):
    """measure -> column index, matched on header text across the header rows.

    Every header row is searched because MHCLG splits labels across two or
    three merged rows and which row carries a given label moves between
    editions.
    """
    found = {}
    for measure, candidates in labels.items():
        hit, ambiguous = None, []
        # Candidates are ordered most-specific first. A candidate that matches
        # exactly one column wins; one that matches several is not resolved by
        # preferring the leftmost, it is abandoned in favour of the next
        # candidate. 2025Q4 gave four A1 columns the prefix 'Assessed as owed
        # a duty', so the old label alone would have picked a column that
        # happened to sort first.
        for cand in candidates:
            hits = {j for r in rows[:8] for j, cell in enumerate(r)
                    if norm_header(cell).startswith(cand)
                    and not is_rate_header(cell)}
            if len(hits) == 1:
                hit = hits.pop()
                break
            if len(hits) > 1:
                ambiguous.append((cand, sorted(hits)))
        if hit is None:
            detail = (f" Ambiguous candidates: {ambiguous}." if ambiguous else "")
            halt(f"{sheet}: no header cell uniquely matches {measure} "
                 f"({candidates}).{detail} The publisher has retitled or "
                 f"restructured a column; resolve it before loading rather "
                 f"than reading by offset.")
        found[measure] = hit
    return found


def num(v):
    """Numeric value, or None. Markers stay absent rather than becoming zero."""
    s = (v or "").strip()
    if s in ("", "-", "..", ":", "*", "n/a", "N/A", "x"):
        return None
    try:
        return int(round(float(s.replace(",", ""))))
    except ValueError:
        return None


def extract(ods, period, source_label):
    sheets = read_sheets(ods, WANTED_SHEETS)
    missing = WANTED_SHEETS - set(sheets)
    if missing:
        halt(f"sheets not found: {sorted(missing)}")

    idx = {s: resolve_columns(sheets[s], COLUMN_LABELS[s], s)
           for s in WANTED_SHEETS}
    for s in sorted(idx):
        print(f"  {s} column indices: "
              + ", ".join(f"{k}={v}" for k, v in sorted(idx[s].items())))

    current, recode = code_resolution()
    resolved_n, unresolved = 0, set()

    by_code = {}
    for sheet in ("A1", "TA1", "A3"):
        for row in sheets[sheet]:
            if not row:
                continue
            published = (row[0] or "").strip()
            if not LA_CODE.fullmatch(published):
                continue
            # Resolve before anything else. Running the orphan check first
            # wastes a gate on a known, predictable condition.
            code = recode.get(published, published)
            if code != published and sheet == "A1":
                resolved_n += 1
            if code not in current:
                unresolved.add(published)
                continue
            rec = by_code.setdefault(code, {"lad24cd": code, "period": period,
                                            "source": source_label})
            if sheet == "A1" and len(row) > 1:
                rec.setdefault("la_name", (row[1] or "").strip())
            for measure, j in idx[sheet].items():
                rec[measure] = num(row[j]) if j < len(row) else None
    if resolved_n:
        print(f"  resolved {resolved_n} published code(s) through la_code_lookup "
              f"(recode only)")
    if unresolved:
        halt(f"{len(unresolved)} published code(s) do not resolve to "
             f"la_boundaries: {sorted(unresolved)}. UNEXPLAINED - establish "
             f"why against an authoritative source before loading.")
    return [by_code[c] for c in sorted(by_code)]


if __name__ == "__main__":
    sys.exit(main())
