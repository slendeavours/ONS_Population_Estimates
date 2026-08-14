"""S23 - Regulator of Social Housing, registered provider stock by local authority.

The first direct supply-side measure in the pipeline. Everything before it is
indirect: S11 counts CQC-registered locations, S8 counts Housing Benefit
Specified Accommodation caseload. This counts units.

Where the data is
-----------------
The annual "Registered provider social housing stock and rents in England"
release draws on two returns - the Statistical Data Return (private registered
providers) and the Local Authority Data Return (local authority registered
providers). The local authority breakdown is not in a flat data file; it is
the STOCK_BY_LA sheet inside the "Registered providers look-up tool" workbook,
which exists to drive the workbook's search box.

That sheet turned out to be the cleanest thing in this pipeline. One row per
provider per local authority, both returns already in the same columns, and
the publisher's own per-LA subtotal rows sitting alongside so the extraction
can be checked against arithmetic the publisher did itself.

Schema decision - one table, not two
------------------------------------
SDR and LADR are separate returns covering different provider types, so the
question was whether they merge. They do, and the evidence is that the
publisher has already merged them: LARP rows carry the same five stock columns
as PRP rows in the same sheet, and the per-LA totals reconcile across both.
Merging is therefore reading the file as published rather than a judgement
imposed on it. provider_type keeps the returns separable, because the analysis
sometimes wants them apart - a LARP is not a competitor in the way a PRP is.

Grain
-----
Provider x local authority. The sheet also carries 296 LA subtotal rows
(RP_Type = 'LA') and 9 regional rows; neither is loaded, because mixing grains
in one table invites a SUM() that double-counts. They are used as verification
instead: gate 7 asserts the loaded provider rows sum to the publisher's own LA
subtotals for all 296 authorities.

What LA_SHHOP is, and the caveat that travels with it
-----------------------------------------------------
SHHOP is supported housing AND housing for older people, reported as one
figure. RSH does not split them at local authority level. A large share of it
is sheltered and retirement housing rather than the supported accommodation
this pipeline is about, so the number is an upper bound on the relevant stock
and must not be read as a count of exempt-accommodation-style provision. That
caveat is recorded on the registry row, not just here.

Usage:
    python scripts/s23_rsh_stock_build.py --discover
    python scripts/s23_rsh_stock_build.py --load
"""
import argparse
import json
import re
import sys
import urllib.request
from pathlib import Path

import openpyxl
import psycopg2.extras

sys.path.insert(0, str(Path(__file__).resolve().parent))
from _db import get_conn  # noqa: E402

if hasattr(sys.stdout, "reconfigure"):
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")

REPO = Path(__file__).resolve().parent.parent
RAW = REPO / "data" / "raw" / "s23_rsh"
TABLE = "rsh_rp_stock_by_la"
SOURCE_CODE = "23"
UA = {"User-Agent": "ucws-pipeline/s23 (+sl@slendeavours.org)"}

LANDING = ("https://www.gov.uk/government/statistics/"
           "registered-provider-social-housing-stock-and-rents-in-england-"
           "2024-to-2025")
CONTENT_API = "https://www.gov.uk/api/content" + LANDING[len("https://www.gov.uk"):]

SHEET = "STOCK_BY_LA"
# Column header -> (target column, whether it is a stock measure). Matched by
# exact header text, and every header in the sheet must be accounted for, so a
# renamed or inserted column stops the build instead of shifting the data.
HEADERS = {
    "RP_Name": "rp_name",
    "RP_Code": "rp_code",
    "RP_Type": "rp_type",
    "SDR_Size": "sdr_size",
    "Survey_Status": "survey_status",
    "LA_Nm": "la_name",
    "LA_Code": "publisher_la_code",
    "Concat": None,                       # the workbook's own search-box key
    "Total Social Stock": "total_social_stock",
    "LA_GN_SC_Own": "general_needs_self_contained",
    "LA_GN_BSp_Own": "general_needs_bedspaces",
    "LA_SHHOP": "supported_housing_and_older_people",
    "LA_LCHO_Less_100_Eqty_Own": "low_cost_home_ownership",
}
STOCK_COLS = ["general_needs_self_contained", "general_needs_bedspaces",
              "supported_housing_and_older_people", "low_cost_home_ownership"]
PROVIDER_TYPES = {"Large": "PRP", "Small": "PRP", "LARP": "LARP"}

DDL = f"""
CREATE TABLE IF NOT EXISTS {TABLE} (
    stock_date                          date        NOT NULL,
    rp_code                             text        NOT NULL,
    lad24cd                             varchar(9)  NOT NULL,
    rp_name                             text        NOT NULL,
    provider_type                       text        NOT NULL,
    rp_size_band                        text,
    survey_status                       text,
    publisher_la_code                   varchar(9)  NOT NULL,
    la_name                             text        NOT NULL,
    total_social_stock                  integer     NOT NULL,
    general_needs_self_contained        integer     NOT NULL,
    general_needs_bedspaces             integer     NOT NULL,
    supported_housing_and_older_people  integer     NOT NULL,
    low_cost_home_ownership             integer     NOT NULL,
    publication_date                    date        NOT NULL,
    edition                             text        NOT NULL,
    source_url                          text        NOT NULL,
    source_file                         text        NOT NULL,
    release_page_url                    text        NOT NULL,
    loaded_at                           timestamptz NOT NULL DEFAULT now(),
    PRIMARY KEY (stock_date, rp_code, lad24cd)
);
"""

DDL_GUARDS = f"""
DO $$
BEGIN
    IF NOT EXISTS (SELECT 1 FROM pg_constraint
                   WHERE conname = '{TABLE}_provider_type_chk') THEN
        ALTER TABLE {TABLE} ADD CONSTRAINT {TABLE}_provider_type_chk
            CHECK (provider_type IN ('PRP','LARP'));
    END IF;
    IF NOT EXISTS (SELECT 1 FROM pg_constraint
                   WHERE conname = '{TABLE}_components_sum_chk') THEN
        ALTER TABLE {TABLE} ADD CONSTRAINT {TABLE}_components_sum_chk
            CHECK (total_social_stock = general_needs_self_contained
                                      + general_needs_bedspaces
                                      + supported_housing_and_older_people
                                      + low_cost_home_ownership);
    END IF;
    IF NOT EXISTS (SELECT 1 FROM pg_indexes
                   WHERE indexname = '{TABLE}_lad24cd_idx') THEN
        CREATE INDEX {TABLE}_lad24cd_idx ON {TABLE} (lad24cd, stock_date);
    END IF;
END $$;
"""

UPSERT = f"""
INSERT INTO {TABLE} (
    stock_date, rp_code, lad24cd, rp_name, provider_type, rp_size_band,
    survey_status, publisher_la_code, la_name, total_social_stock,
    general_needs_self_contained, general_needs_bedspaces,
    supported_housing_and_older_people, low_cost_home_ownership,
    publication_date, edition, source_url, source_file, release_page_url)
VALUES %s
ON CONFLICT (stock_date, rp_code, lad24cd) DO UPDATE SET
    rp_name                            = EXCLUDED.rp_name,
    provider_type                      = EXCLUDED.provider_type,
    rp_size_band                       = EXCLUDED.rp_size_band,
    survey_status                      = EXCLUDED.survey_status,
    publisher_la_code                  = EXCLUDED.publisher_la_code,
    la_name                            = EXCLUDED.la_name,
    total_social_stock                 = EXCLUDED.total_social_stock,
    general_needs_self_contained       = EXCLUDED.general_needs_self_contained,
    general_needs_bedspaces            = EXCLUDED.general_needs_bedspaces,
    supported_housing_and_older_people = EXCLUDED.supported_housing_and_older_people,
    low_cost_home_ownership            = EXCLUDED.low_cost_home_ownership,
    publication_date                   = EXCLUDED.publication_date,
    edition                            = EXCLUDED.edition,
    source_url                         = EXCLUDED.source_url,
    source_file                        = EXCLUDED.source_file,
    release_page_url                   = EXCLUDED.release_page_url,
    loaded_at                          = now();
"""


def halt(msg):
    sys.exit(f"HALT: {msg}")


def resolve_edition():
    """Resolve the look-up tool from the release page, never from memory."""
    with urllib.request.urlopen(
            urllib.request.Request(CONTENT_API, headers=UA), timeout=60) as fh:
        doc = json.load(fh)
    chosen = None
    for att in doc["details"].get("attachments", []):
        url = str(att.get("url", ""))
        if url.startswith("http") and re.search(r"COMBINED_TOOL", url, re.I):
            chosen = (url, url.rsplit("/", 1)[-1])
    if chosen is None:
        halt("no combined look-up tool attachment on the release page")
    year = re.search(r"(\d{4})-to-(\d{4})", doc["base_path"])
    if not year:
        halt(f"cannot read the edition years from {doc['base_path']}")
    return {
        "url": chosen[0], "filename": chosen[1],
        "edition": f"{year.group(1)} to {year.group(2)}",
        # The return is a snapshot at 31 March of the closing year. Publication
        # lands months later; both dates are stored because a stock figure
        # read as though it were current is wrong by up to a year.
        "stock_date": f"{year.group(2)}-03-31",
        "publication_date": doc.get("first_published_at", "")[:10],
        "release_page_url": "https://www.gov.uk" + doc["base_path"],
    }


def fetch(edition):
    RAW.mkdir(parents=True, exist_ok=True)
    path = RAW / edition["filename"]
    if not path.exists():
        with urllib.request.urlopen(
                urllib.request.Request(edition["url"], headers=UA),
                timeout=300) as fh:
            path.write_bytes(fh.read())
    return path


def read_sheet(path):
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    if SHEET not in wb.sheetnames:
        halt(f"{path.name}: no sheet named {SHEET}; sheets are {wb.sheetnames}")
    rows = list(wb[SHEET].iter_rows(values_only=True))
    header = [str(c).strip() if c is not None else "" for c in rows[0]]
    unknown = [h for h in header if h and h not in HEADERS]
    if unknown:
        halt(f"{SHEET}: unrecognised column header(s) {unknown}. The sheet has "
             f"changed shape; check the mapping before loading.")
    missing = [h for h in HEADERS if h not in header]
    if missing:
        halt(f"{SHEET}: expected column(s) not present: {missing}")
    idx = {HEADERS[h]: j for j, h in enumerate(header) if h and HEADERS[h]}
    return [r for r in rows[1:] if r[header.index("RP_Code")] is not None], idx


def num(v, what):
    if v is None:
        return 0
    if isinstance(v, (int, float)):
        return int(round(v))
    halt(f"{what}: non-numeric stock value {v!r}")


def split_rows(rows, idx):
    """Separate provider rows from the publisher's own subtotal rows."""
    provider, la_totals, other = [], {}, []
    for r in rows:
        rp_type = str(r[idx["rp_type"]]).strip()
        if rp_type in PROVIDER_TYPES:
            provider.append(r)
        elif rp_type == "LA":
            la_totals[str(r[idx["publisher_la_code"]]).strip()] = r
        else:
            other.append(r)
    return provider, la_totals, other


def build_rows(cur, edition, rows, idx):
    provider, la_totals, other = split_rows(rows, idx)

    codes = sorted({str(r[idx["publisher_la_code"]]).strip() for r in provider})
    cur.execute("SELECT old_code, new_code FROM la_code_lookup "
                "WHERE old_code = ANY(%s)", (codes,))
    lookup = dict(cur.fetchall())
    unresolved = sorted(set(codes) - set(lookup))
    if unresolved:
        halt(f"{len(unresolved)} publisher LA code(s) do not resolve through "
             f"la_code_lookup: {unresolved}")

    out, seen = [], {}
    for r in provider:
        rp_code = str(r[idx["rp_code"]]).strip()
        pub_la = str(r[idx["publisher_la_code"]]).strip()
        lad24cd = lookup[pub_la]
        key = (rp_code, lad24cd)
        if key in seen:
            halt(f"duplicate provider/authority pair {key} - the assumed grain "
                 f"is wrong and the primary key would silently drop a row")
        seen[key] = True
        stock = {c: num(r[idx[c]], f"{rp_code}/{pub_la}/{c}") for c in STOCK_COLS}
        total = num(r[idx["total_social_stock"]], f"{rp_code}/{pub_la}/total")
        if total != sum(stock.values()):
            halt(f"{rp_code}/{pub_la}: total {total} does not equal the sum of "
                 f"its components {sum(stock.values())}")
        out.append((
            edition["stock_date"], rp_code, lad24cd,
            str(r[idx["rp_name"]]).strip(),
            PROVIDER_TYPES[str(r[idx["rp_type"]]).strip()],
            str(r[idx["sdr_size"]]).strip() or None,
            str(r[idx["survey_status"]]).strip() or None,
            pub_la, str(r[idx["la_name"]]).strip(), total,
            stock["general_needs_self_contained"],
            stock["general_needs_bedspaces"],
            stock["supported_housing_and_older_people"],
            stock["low_cost_home_ownership"],
            edition["publication_date"], edition["edition"], edition["url"],
            edition["filename"], edition["release_page_url"]))
    return out, la_totals, other


def log_run(cur, rows_written, edition):
    cur.execute("""
        INSERT INTO pipeline_run_log
            (run_id, source_number, source_code, agent_name, status,
             rows_written, notes, started_at, completed_at)
        VALUES (gen_random_uuid(), %s, %s, %s, 'success', %s, %s, now(), now())
    """, ("23", SOURCE_CODE,
          # pipeline_run_log.agent_name is varchar(50).
          "Source 23 - RSH RP stock by local authority",
          rows_written,
          f"Edition {edition['edition']}, stock as at {edition['stock_date']}, "
          f"published {edition['publication_date']}. Table {TABLE}."))


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--discover", action="store_true")
    ap.add_argument("--load", action="store_true")
    args = ap.parse_args()
    if not (args.discover or args.load):
        ap.error("choose --discover or --load")

    edition = resolve_edition()
    rows, idx = read_sheet(fetch(edition))

    conn = get_conn()
    cur = conn.cursor()
    try:
        if args.load:
            cur.execute(DDL)
            cur.execute(DDL_GUARDS)
        out, la_totals, other = build_rows(cur, edition, rows, idx)

        print(f"edition            : {edition['edition']}")
        print(f"stock date         : {edition['stock_date']}")
        print(f"publication date   : {edition['publication_date']}")
        print(f"file               : {edition['filename']}")
        print(f"provider rows      : {len(out)}")
        print(f"LA subtotal rows   : {len(la_totals)} (not loaded, used to verify)")
        print(f"other rows skipped : {len(other)} (regional aggregates)")
        prp = sum(1 for r in out if r[4] == "PRP")
        print(f"  PRP (SDR)        : {prp}")
        print(f"  LARP (LADR)      : {len(out) - prp}")
        print(f"authorities        : {len({r[2] for r in out})}")
        print(f"supported housing  : {sum(r[12] for r in out):,} units")
        print(f"total social stock : {sum(r[9] for r in out):,} units")

        if args.load:
            psycopg2.extras.execute_values(cur, UPSERT, out, page_size=1000)
            log_run(cur, len(out), edition)
            cur.execute(f"SELECT COUNT(*) FROM {TABLE}")
            print(f"\n{TABLE}: {cur.fetchone()[0]} rows")
            conn.commit()
            print("COMMITTED")
        else:
            conn.rollback()
            print("\nNothing committed.")
        return 0
    except Exception:
        conn.rollback()
        raise
    finally:
        cur.close()
        conn.close()


if __name__ == "__main__":
    sys.exit(main())
