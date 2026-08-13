"""S3 — refresh la_population to the latest ONS mid-year estimates.

Why now: pip_rate_per_1000 was Apr-26 PIP claimants over a mid-2024
population base. Two years of drift in the denominator of a rate that goes
in front of a council or an NHS commissioner. ONS published the mid-2025
estimates on 29 July 2026, so this is a refresh rather than a wait.

No file URL is hardcoded. The edition is resolved at run time from the ONS
dataset landing page, which also re-confirms the publication date
independently of anything written in this repository.

Two traps were carried into this build. One held, one did not:

  HELD     The release covers 318 England-and-Wales local authorities.
           England is filtered on the code prefix, never assumed to be the
           whole file. 296 England rows, 22 Welsh unitary authorities.

  DID NOT  The expectation was that a release published after 1 April 2025
           would use the recoded Barnsley and Sheffield codes E08000038 and
           E08000039. It does not. The edition is explicitly "2023 local
           authority boundaries", so it publishes E08000016 and E08000019 —
           the same codes la_boundaries carries. Verified in the file rather
           than assumed in either direction. The predictor is the edition's
           declared boundary vintage, not the publication date.

la_population gains a composite key so mid-2024 and mid-2025 coexist. That
makes the W1 node 5 join fan out unless it is pinned to the latest year;
see scripts/s3_w1_wire.py.
"""
import datetime
import json
import re
import sys
import urllib.request
from pathlib import Path

import openpyxl
import psycopg2.extras

sys.path.insert(0, str(Path(__file__).resolve().parent))
from _db import get_conn  # noqa: E402

REPO = Path(__file__).resolve().parent.parent
RAW_DIR = REPO / "data" / "raw" / "s3_mye"
REPORT_DIR = REPO / "build_reports"

ONS = "https://www.ons.gov.uk"
LANDING = ("/peoplepopulationandcommunity/populationandmigration/"
           "populationestimates/datasets/estimatesofthepopulationfor"
           "englandandwales")
UA = {"User-Agent": "ucws-pipeline/s3 (+sl@slendeavours.org)"}

SHEET = "MYE2 - Persons"
HEADER_ROW = 8
ENGLAND_PREFIXES = ("06", "07", "08", "09")
SOURCE_NUMBER = "3"
AGENT_NAME = "Source 3 - ONS Mid-Year Population Estimates"


def halt(msg):
    sys.exit(f"HALT: {msg}")


def _get_json(path):
    req = urllib.request.Request(ONS + path + "/data", headers=UA)
    with urllib.request.urlopen(req, timeout=120) as r:
        return json.loads(r.read().decode("utf-8"))


def discover():
    """Latest mid-year edition, resolved from the landing page."""
    page = _get_json(LANDING)
    editions = []
    for d in page.get("datasets", []):
        uri = d.get("uri", "")
        m = re.search(r"/mid(\d{4})(\d{4})localauthorityboundaries$", uri)
        if m:
            editions.append((int(m.group(1)), int(m.group(2)), uri))
    if not editions:
        halt(f"no 'mid<year><boundary>localauthorityboundaries' edition found "
             f"on {LANDING} — landing page structure has changed")
    editions.sort(reverse=True)
    mid_year, boundary_year, uri = editions[0]

    ed = _get_json(uri)
    desc = ed.get("description", {})
    downloads = ed.get("downloads", [])
    if not downloads or not downloads[0].get("file"):
        halt(f"edition {uri} has no download file attached")
    filename = downloads[0]["file"]
    url = f"{ONS}/file?uri={uri}/{filename}"

    dest = RAW_DIR / filename
    if not dest.exists() or dest.stat().st_size == 0:
        dest.parent.mkdir(parents=True, exist_ok=True)
        req = urllib.request.Request(url, headers=UA)
        with urllib.request.urlopen(req, timeout=600) as r, open(dest, "wb") as f:
            f.write(r.read())

    return {
        "mid_year": mid_year,
        "boundary_year": boundary_year,
        "edition": desc.get("edition"),
        "release_date": (desc.get("releaseDate") or "")[:10],
        "landing_page": ONS + LANDING,
        "edition_page": ONS + uri,
        "url": url,
        "filename": filename,
        "path": dest,
        "size": dest.stat().st_size,
    }


def extract(src):
    wb = openpyxl.load_workbook(src["path"], read_only=True, data_only=True)
    if SHEET not in wb.sheetnames:
        halt(f"sheet '{SHEET}' not found — workbook structure has changed. "
             f"Sheets present: {wb.sheetnames}")
    ws = wb[SHEET]
    rows = list(ws.iter_rows(min_row=HEADER_ROW, values_only=True))
    header, data = rows[0], rows[1:]
    if [str(h).strip() for h in header[:4]] != ["Code", "Name", "Geography",
                                                "All ages"]:
        halt(f"unexpected headers on {SHEET}: {header[:4]} — structure has "
             "changed")

    england_row = next((r for r in data if str(r[0]) == "E92000001"), None)
    if england_row is None:
        halt("England country row (E92000001) not found — cannot reconcile")

    records, wales = [], 0
    for r in data:
        code = str(r[0]).strip() if r[0] else ""
        if code.startswith("W"):
            wales += 1
            continue
        # England local authorities only. E10/E11 counties, E12 regions and
        # E92 country rows are aggregates and would double count.
        if not (code.startswith("E") and code[1:3] in ENGLAND_PREFIXES):
            continue
        pop = r[3]
        if not isinstance(pop, (int, float)):
            halt(f"{code} has a non-numeric All ages value: {pop!r}")
        records.append({"lad24cd": code, "la_name": str(r[1]).strip(),
                        "population": int(pop),
                        "reference_year": src["mid_year"]})
    wb.close()
    return records, int(england_row[3]), wales


MIGRATION = """
DO $$
BEGIN
    IF NOT EXISTS (
        SELECT 1 FROM information_schema.columns
         WHERE table_name = 'la_population'
           AND column_name = 'reference_year') THEN
        RAISE EXCEPTION 'la_population has no reference_year column';
    END IF;

    -- Widen the key so successive vintages coexist instead of overwriting.
    IF EXISTS (
        SELECT 1 FROM pg_constraint
         WHERE conrelid = 'la_population'::regclass
           AND conname = 'la_population_pkey'
           AND pg_get_constraintdef(oid) = 'PRIMARY KEY (lad24cd)') THEN
        ALTER TABLE la_population DROP CONSTRAINT la_population_pkey;
        ALTER TABLE la_population
            ADD CONSTRAINT la_population_pkey
            PRIMARY KEY (lad24cd, reference_year);
        RAISE NOTICE 'la_population key widened to (lad24cd, reference_year)';
    END IF;

    ALTER TABLE la_population ALTER COLUMN reference_year SET NOT NULL;
END $$;
"""

UPSERT = """
INSERT INTO la_population (lad24cd, la_name, population, reference_year,
                           loaded_at)
VALUES (%(lad24cd)s, %(la_name)s, %(population)s, %(reference_year)s, now())
ON CONFLICT (lad24cd, reference_year) DO UPDATE SET
    la_name    = EXCLUDED.la_name,
    population = EXCLUDED.population,
    loaded_at  = now()
"""


def main():
    REPORT_DIR.mkdir(parents=True, exist_ok=True)
    started = datetime.datetime.now(datetime.timezone.utc)

    print("Phase 1: discovering the latest mid-year edition")
    src = discover()
    print(f"  edition      : {src['edition']}")
    print(f"  mid-year     : {src['mid_year']}   "
          f"boundary vintage: {src['boundary_year']}")
    print(f"  release date : {src['release_date']}")
    print(f"  file         : {src['filename']} ({src['size']:,} bytes)")

    records, england_published, wales = extract(src)
    print(f"  England LA rows: {len(records)}   Welsh rows skipped: {wales}")

    conn = get_conn()
    conn.autocommit = False
    cur = conn.cursor()

    cur.execute("SELECT COUNT(*), MIN(reference_year), MAX(reference_year) "
                "FROM la_population")
    before = cur.fetchone()
    print(f"  la_population before: {before[0]} rows, years "
          f"{before[1]}-{before[2]}")

    print("Phase 2: migrate key, load")
    cur.execute(MIGRATION)
    psycopg2.extras.execute_batch(cur, UPSERT, records, page_size=200)

    gates = []

    def gate(n, name, ok, detail):
        gates.append(dict(num=n, name=name, ok=bool(ok), detail=detail))
        print(f"  gate {n} {'PASS' if ok else 'FAIL'}: {name} — {detail}")

    yr = src["mid_year"]
    cur.execute("SELECT COUNT(*) FROM la_population WHERE reference_year=%s",
                (yr,))
    n = cur.fetchone()[0]
    gate(1, "296 England rows for the new reference year", n == 296,
         f"{n} rows at reference_year {yr}")

    cur.execute("""SELECT COUNT(*) FROM la_population p
                    WHERE p.reference_year=%s
                      AND NOT EXISTS (SELECT 1 FROM la_boundaries b
                                       WHERE b.lad24cd=p.lad24cd)""", (yr,))
    orph = cur.fetchone()[0]
    gate(2, "every lad24cd exists in la_boundaries", orph == 0,
         f"{orph} orphan codes")

    cur.execute("SELECT SUM(population) FROM la_population "
                "WHERE reference_year=%s", (yr,))
    loaded_total = cur.fetchone()[0]
    diff = abs(loaded_total - england_published) / england_published * 100
    gate(3, "England total reconciles to the publisher's own England row",
         diff == 0,
         f"loaded {loaded_total:,} vs published {england_published:,} "
         f"({diff:.4f}%)")

    cur.execute("SELECT COUNT(*) FROM la_population WHERE population <= 0")
    bad = cur.fetchone()[0]
    gate(4, "no non-positive populations", bad == 0, f"{bad} rows")

    cur.execute("SELECT COUNT(*) FROM la_population WHERE reference_year=%s",
                (before[2],))
    kept = cur.fetchone()[0]
    gate(5, "the previous vintage is retained, not overwritten",
         kept == 296 or before[2] == yr,
         f"{kept} rows still at reference_year {before[2]}")

    if any(not g["ok"] for g in gates):
        conn.rollback()
        conn.close()
        sys.exit("HALT: hard gate failed — rolled back, la_population left in "
                 "its pre-load state.")
    conn.commit()

    print("Gate 6: idempotency — re-running the load")
    cur.execute("SELECT COUNT(*), SUM(population) FROM la_population")
    snap_before = cur.fetchone()
    psycopg2.extras.execute_batch(cur, UPSERT, records, page_size=200)
    conn.commit()
    cur.execute("SELECT COUNT(*), SUM(population) FROM la_population")
    snap_after = cur.fetchone()
    same = snap_before == snap_after
    gate(6, "second load changes nothing", same,
         f"{snap_before} -> {snap_after}")
    if not same:
        conn.close()
        sys.exit("HALT: idempotency gate failed.")

    notes = (
        f"ONS mid-{yr} population estimates, {src['edition']}, released "
        f"{src['release_date']} ({src['url']}). 296 England local "
        f"authorities loaded; {wales} Welsh rows in the file were filtered "
        "out on code prefix rather than assumed absent. Published on "
        f"{src['boundary_year']} local authority boundaries, so Barnsley and "
        "Sheffield appear as E08000016/E08000019 and needed no recode "
        "resolution despite the release postdating the April 2025 recode. "
        f"la_population key widened to (lad24cd, reference_year); the "
        f"mid-{before[2]} vintage is retained. England total "
        f"{loaded_total:,}, reconciling exactly to the publisher's England "
        "row."
    )
    cur.execute("""
        INSERT INTO pipeline_run_log
            (run_id, agent_name, source_number, status, rows_written,
             error_message, started_at, completed_at, duration_ms, notes)
        VALUES (gen_random_uuid(), %s, %s, 'complete', %s, NULL, %s, now(),
                NULL, %s) RETURNING id
    """, (AGENT_NAME, SOURCE_NUMBER, len(records), started, notes))
    log_id = cur.fetchone()[0]
    conn.commit()

    state = dict(src, path=str(src["path"]), gates=gates,
                 england_published=england_published,
                 wales_rows_skipped=wales, records=len(records),
                 previous_year=before[2], log_id=log_id,
                 started_at=started.isoformat())
    (REPORT_DIR / "s3_mye_state.json").write_text(
        json.dumps(state, indent=2, default=str), encoding="utf-8")
    print(f"\nlogged to pipeline_run_log id {log_id}")
    print("wrote build_reports/s3_mye_state.json")
    conn.close()


if __name__ == "__main__":
    main()
