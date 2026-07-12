"""
s18_pipr_inspect.py — S18 Phase 2: Inspect the PIPR workbook structure.

Purpose : List every worksheet, locate header rows, print column headers,
          row counts, and sample rows so the structure document can record
          facts rather than assumptions. No transformation happens here.
Inputs  : data/raw/pipr_<edition>.xlsx (path passed as argv[1])
Outputs : stdout report only.
"""
import sys
from pathlib import Path

import openpyxl


def main(path):
    wb = openpyxl.load_workbook(path, read_only=True)
    for name in wb.sheetnames:
        ws = wb[name]
        print(f"\n=== SHEET: {name!r}  dims={ws.calculate_dimension()} "
              f"max_row={ws.max_row} max_col={ws.max_column}")
        # print first 8 rows, first 12 cells each, to find the header row
        for i, row in enumerate(ws.iter_rows(min_row=1, max_row=8,
                                             max_col=12, values_only=True), 1):
            cells = [str(c)[:28] if c is not None else "" for c in row]
            print(f"  r{i}: {cells}")
    wb.close()


if __name__ == "__main__":
    main(sys.argv[1] if len(sys.argv) > 1 else
         str(Path(__file__).resolve().parent.parent / "data/raw/pipr_17june2026.xlsx"))
