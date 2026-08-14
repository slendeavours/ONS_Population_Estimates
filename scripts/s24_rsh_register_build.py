"""S24 - RSH register of registered providers, and regulatory judgements.

Purpose is risk management, not analysis. The income route runs through a
registered provider partner, so a downgraded governance grade, an enforcement
notice or a de-registration is a material event for the business. The point of
holding this is to be told rather than to find out.

Deliberately not wired into Workflow 1
--------------------------------------
RSH does not publish provider addresses or contact details, so this source has
no local authority geography and cannot be apportioned to one. It must not be
joined into staging_la_signals and must not appear as a map layer. That is a
design decision, recorded here and on the registry row so it is not wired in
later by reflex: a provider's registered office is not where its stock is, and
inventing a geography from one would be worse than having none.

What is published, and what discovery established
-------------------------------------------------
Two machine-readable files, not one:

  * The monthly register snapshot - an xlsx of every current provider, with
    registration number, name, registration date, designation and corporate
    form. Published around the middle of each month. It carries only the
    current month; there is no archive of past snapshots, so history exists
    only if this table keeps it.

  * The regulatory judgements and notices table - an xlsx with consumer,
    governance and viability grades per provider, each with its own grade
    date and a change description, plus a separate sheet of enforcement
    notices. The prompt for this build allowed for judgements being available
    only as individual documents, in which case they were to be recorded as a
    limitation. They are not: this is a proper table and it is loaded.

De-registration is inferred, and that is stated rather than hidden
-----------------------------------------------------------------
The snapshot lists current providers only. A de-registration is therefore
visible as an absence, not as an event with a date. The register table stores
one row per provider per snapshot, so "what changed since last month" is
answerable from the table itself by comparing two snapshot dates - which is
the whole reason it is stored as snapshots rather than as a current-state
table that would overwrite the evidence.

Usage:
    python scripts/s24_rsh_register_build.py --discover
    python scripts/s24_rsh_register_build.py --load
"""
import argparse
import datetime as dt
import json
import re
import sys
import urllib.request
from pathlib import Path

import openpyxl
import psycopg2.extras

sys.path.insert(0, str(Path(__file__).resolve().parent))
from _db import get_conn  # noqa: E402

if hasattr(sys.stdout, "reconfigure"):
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")

REPO = Path(__file__).resolve().parent.parent
RAW = REPO / "data" / "raw" / "s24_rsh"
REGISTER_TABLE = "rsh_registered_providers"
JUDGEMENT_TABLE = "rsh_regulatory_judgements"
NOTICE_TABLE = "rsh_enforcement_notices"
SOURCE_CODE = "24"
UA = {"User-Agent": "ucws-pipeline/s24 (+sl@slendeavours.org)"}

REGISTER_PAGE = ("https://www.gov.uk/government/publications/"
                 "registered-providers-of-social-housing")
JUDGEMENT_PAGE = ("https://www.gov.uk/government/publications/"
                  "regulatory-judgements-and-enforcement-notices")
API = "https://www.gov.uk/api/content"

REGISTER_HEADERS = ["Organisation name", "Registration number",
                    "Registration date", "Designation", "Corporate form",
                    "Notes"]
JUDGEMENT_HEADERS = [
    "Reg Code", "Landlord", "Landlord Type",
    "Name and Reg Code Change Details", "Other landlords included in the judgement",
    "Status", "Consumer grade", "Consumer Grade Change", "Consumer Grade Date",
    "Governance Grade", "Governance Grade Change", "Governance Date",
    "Viability Grade", "Viability Grade Change", "Viability Grade Date",
    "Rent", "Rent Date", "Rent Change", "Type of Publication",
    "Publication Date", "Engagement Process"]
NOTICE_HEADERS = [
    "Reg Code", "Provider", "Name and Reg Code Change Details",
    "Other providers included in the notice", "Status", "Type of Publication",
    "Publication Date", "Route", "Explanation", "Date of Enforcement Notice"]

DDL = f"""
CREATE TABLE IF NOT EXISTS {REGISTER_TABLE} (
    snapshot_date       date        NOT NULL,
    registration_number text        NOT NULL,
    organisation_name   text        NOT NULL,
    registration_date   date,
    designation         text,
    corporate_form      text,
    notes               text,
    source_url          text        NOT NULL,
    source_file         text        NOT NULL,
    release_page_url    text        NOT NULL,
    loaded_at           timestamptz NOT NULL DEFAULT now(),
    PRIMARY KEY (snapshot_date, registration_number)
);

CREATE TABLE IF NOT EXISTS {JUDGEMENT_TABLE} (
    registration_number   text        NOT NULL,
    publication_date      date        NOT NULL,
    landlord_name         text        NOT NULL,
    landlord_type         text,
    status                text,
    consumer_grade        text,
    consumer_grade_change text,
    consumer_grade_date   date,
    governance_grade      text,
    governance_grade_change text,
    governance_grade_date date,
    viability_grade       text,
    viability_grade_change text,
    viability_grade_date  date,
    rent_grade            text,
    rent_grade_change     text,
    rent_grade_date       date,
    publication_type      text,
    engagement_process    text,
    name_or_code_change   text,
    other_landlords       text,
    edition_date          date        NOT NULL,
    source_url            text        NOT NULL,
    source_file           text        NOT NULL,
    release_page_url      text        NOT NULL,
    loaded_at             timestamptz NOT NULL DEFAULT now(),
    PRIMARY KEY (registration_number, publication_date)
);

CREATE TABLE IF NOT EXISTS {NOTICE_TABLE} (
    registration_number  text        NOT NULL,
    publication_date     date        NOT NULL,
    provider_name        text        NOT NULL,
    status               text,
    publication_type     text,
    route                text,
    explanation          text,
    notice_date          date,
    name_or_code_change  text,
    other_providers      text,
    edition_date         date        NOT NULL,
    source_url           text        NOT NULL,
    source_file          text        NOT NULL,
    release_page_url     text        NOT NULL,
    loaded_at            timestamptz NOT NULL DEFAULT now(),
    PRIMARY KEY (registration_number, publication_date)
);
"""

DDL_GUARDS = f"""
DO $$
BEGIN
    IF NOT EXISTS (SELECT 1 FROM pg_indexes
                   WHERE indexname = '{REGISTER_TABLE}_regnum_idx') THEN
        CREATE INDEX {REGISTER_TABLE}_regnum_idx
            ON {REGISTER_TABLE} (registration_number, snapshot_date);
    END IF;
END $$;
"""

REGISTER_UPSERT = f"""
INSERT INTO {REGISTER_TABLE} (
    snapshot_date, registration_number, organisation_name, registration_date,
    designation, corporate_form, notes, source_url, source_file,
    release_page_url)
VALUES %s
ON CONFLICT (snapshot_date, registration_number) DO UPDATE SET
    organisation_name = EXCLUDED.organisation_name,
    registration_date = EXCLUDED.registration_date,
    designation       = EXCLUDED.designation,
    corporate_form    = EXCLUDED.corporate_form,
    notes             = EXCLUDED.notes,
    source_url        = EXCLUDED.source_url,
    source_file       = EXCLUDED.source_file,
    release_page_url  = EXCLUDED.release_page_url,
    loaded_at         = now();
"""

JUDGEMENT_UPSERT = f"""
INSERT INTO {JUDGEMENT_TABLE} (
    registration_number, publication_date, landlord_name, landlord_type,
    status, consumer_grade, consumer_grade_change, consumer_grade_date,
    governance_grade, governance_grade_change, governance_grade_date,
    viability_grade, viability_grade_change, viability_grade_date,
    rent_grade, rent_grade_change, rent_grade_date, publication_type,
    engagement_process, name_or_code_change, other_landlords, edition_date,
    source_url, source_file, release_page_url)
VALUES %s
ON CONFLICT (registration_number, publication_date) DO UPDATE SET
    landlord_name          = EXCLUDED.landlord_name,
    landlord_type          = EXCLUDED.landlord_type,
    status                 = EXCLUDED.status,
    consumer_grade         = EXCLUDED.consumer_grade,
    consumer_grade_change  = EXCLUDED.consumer_grade_change,
    consumer_grade_date    = EXCLUDED.consumer_grade_date,
    governance_grade       = EXCLUDED.governance_grade,
    governance_grade_change = EXCLUDED.governance_grade_change,
    governance_grade_date  = EXCLUDED.governance_grade_date,
    viability_grade        = EXCLUDED.viability_grade,
    viability_grade_change = EXCLUDED.viability_grade_change,
    viability_grade_date   = EXCLUDED.viability_grade_date,
    rent_grade             = EXCLUDED.rent_grade,
    rent_grade_change      = EXCLUDED.rent_grade_change,
    rent_grade_date        = EXCLUDED.rent_grade_date,
    publication_type       = EXCLUDED.publication_type,
    engagement_process     = EXCLUDED.engagement_process,
    name_or_code_change    = EXCLUDED.name_or_code_change,
    other_landlords        = EXCLUDED.other_landlords,
    edition_date           = EXCLUDED.edition_date,
    source_url             = EXCLUDED.source_url,
    source_file            = EXCLUDED.source_file,
    release_page_url       = EXCLUDED.release_page_url,
    loaded_at              = now();
"""

NOTICE_UPSERT = f"""
INSERT INTO {NOTICE_TABLE} (
    registration_number, publication_date, provider_name, status,
    publication_type, route, explanation, notice_date, name_or_code_change,
    other_providers, edition_date, source_url, source_file, release_page_url)
VALUES %s
ON CONFLICT (registration_number, publication_date) DO UPDATE SET
    provider_name       = EXCLUDED.provider_name,
    status              = EXCLUDED.status,
    publication_type    = EXCLUDED.publication_type,
    route               = EXCLUDED.route,
    explanation         = EXCLUDED.explanation,
    notice_date         = EXCLUDED.notice_date,
    name_or_code_change = EXCLUDED.name_or_code_change,
    other_providers     = EXCLUDED.other_providers,
    edition_date        = EXCLUDED.edition_date,
    source_url          = EXCLUDED.source_url,
    source_file         = EXCLUDED.source_file,
    release_page_url    = EXCLUDED.release_page_url,
    loaded_at           = now();
"""


def halt(msg):
    sys.exit(f"HALT: {msg}")


def _api(page_url):
    with urllib.request.urlopen(
            urllib.request.Request(API + page_url[len("https://www.gov.uk"):],
                                   headers=UA), timeout=60) as fh:
        return json.load(fh)


def resolve_register():
    """The current monthly snapshot. Its date comes from the attachment title,
    not from today - the file is published mid-month and a run on any later day
    must still record the snapshot the publisher named."""
    doc = _api(REGISTER_PAGE)
    for att in doc["details"].get("attachments", []):
        url = str(att.get("url", ""))
        if not url.startswith("http") or not url.lower().endswith(".xlsx"):
            continue
        title = att.get("title", "")
        m = re.search(r"(\d{1,2})\s+(\w+)\s+(\d{4})", title)
        if not m:
            continue
        try:
            snapshot = dt.datetime.strptime(
                f"{m.group(1)} {m.group(2)} {m.group(3)}", "%d %B %Y").date()
        except ValueError:
            continue
        return {"url": url, "filename": url.rsplit("/", 1)[-1],
                "snapshot_date": snapshot.isoformat(),
                "release_page_url": REGISTER_PAGE}
    halt("no dated register spreadsheet found on the register page")


def resolve_judgements():
    doc = _api(JUDGEMENT_PAGE)
    for att in doc["details"].get("attachments", []):
        url = str(att.get("url", ""))
        if url.startswith("http") and url.lower().endswith(".xlsx"):
            name = url.rsplit("/", 1)[-1]
            m = re.match(r"(\d{4})(\d{2})(\d{2})", name)
            edition = (f"{m.group(1)}-{m.group(2)}-{m.group(3)}" if m
                       else doc.get("public_updated_at", "")[:10])
            return {"url": url, "filename": name, "edition_date": edition,
                    "release_page_url": JUDGEMENT_PAGE}
    halt("no judgements spreadsheet found on the judgements page")


def fetch(spec):
    RAW.mkdir(parents=True, exist_ok=True)
    path = RAW / spec["filename"]
    if not path.exists():
        with urllib.request.urlopen(
                urllib.request.Request(spec["url"], headers=UA),
                timeout=180) as fh:
            path.write_bytes(fh.read())
    return path


def sheet_rows(path, wanted_headers, sheet_hint):
    """Read a sheet, asserting its header is exactly what is expected.

    The register workbook also carries a hidden sheet holding what looks like
    a stray account token from the publisher's own tooling. It is not read and
    nothing from it is stored.
    """
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    target = None
    for ws in wb.worksheets:
        if sheet_hint.lower() in ws.title.lower():
            target = ws
            break
    if target is None:
        halt(f"{path.name}: no sheet matching {sheet_hint!r}; "
             f"sheets are {wb.sheetnames}")
    rows = list(target.iter_rows(values_only=True))
    header = [str(c).strip() if c is not None else "" for c in rows[0]]
    header = header[:len(wanted_headers)]
    if header != wanted_headers:
        halt(f"{path.name}/{target.title}: header changed.\n"
             f"  expected: {wanted_headers}\n  found:    {header}")
    return [r for r in rows[1:] if r[0] is not None and str(r[0]).strip()]


def text(v):
    if v is None:
        return None
    s = str(v).strip()
    # '-' is the publisher's marker for a grade that has not been assessed. It
    # is stored as NULL, and the paired "Change" column carries the reason, so
    # "not assessed" stays distinguishable from a missing cell.
    return None if s in ("", "-", "None") else s


def date(v):
    if v is None:
        return None
    if isinstance(v, dt.datetime):
        return v.date().isoformat()
    if isinstance(v, dt.date):
        return v.isoformat()
    # Some cells survive as raw Excel serials rather than typed dates, because
    # the publisher's workbook leaves them unformatted. The 1899-12-30 epoch is
    # Excel's, and the range guard stops a stray count being read as a date.
    if isinstance(v, (int, float)) and not isinstance(v, bool):
        if not 20000 <= v <= 60000:
            halt(f"numeric value {v!r} in a date column is outside any "
                 f"plausible Excel date serial range")
        return (dt.date(1899, 12, 30) + dt.timedelta(days=int(v))).isoformat()
    s = str(v).strip()
    if not s or s == "-":
        return None
    for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d", "%d/%m/%Y"):
        try:
            return dt.datetime.strptime(s, fmt).date().isoformat()
        except ValueError:
            continue
    halt(f"unparseable date {v!r}")


def build_register(spec, rows):
    out, seen = [], set()
    for r in rows:
        num = str(r[1]).strip()
        if num in seen:
            halt(f"duplicate registration number {num} in one snapshot")
        seen.add(num)
        out.append((spec["snapshot_date"], num, str(r[0]).strip(), date(r[2]),
                    text(r[3]), text(r[4]), text(r[5]) if len(r) > 5 else None,
                    spec["url"], spec["filename"], spec["release_page_url"]))
    return out


def build_judgements(spec, rows):
    out, seen = [], set()
    for r in rows:
        num, pub = str(r[0]).strip(), date(r[19])
        if pub is None:
            halt(f"judgement for {num} has no publication date")
        if (num, pub) in seen:
            halt(f"duplicate judgement key {(num, pub)}")
        seen.add((num, pub))
        out.append((num, pub, str(r[1]).strip(), text(r[2]), text(r[5]),
                    text(r[6]), text(r[7]), date(r[8]),
                    text(r[9]), text(r[10]), date(r[11]),
                    text(r[12]), text(r[13]), date(r[14]),
                    text(r[15]), text(r[17]), date(r[16]),
                    text(r[18]), text(r[20]), text(r[3]), text(r[4]),
                    spec["edition_date"], spec["url"], spec["filename"],
                    spec["release_page_url"]))
    return out


def build_notices(spec, rows):
    out, seen = [], set()
    for r in rows:
        num, pub = str(r[0]).strip(), date(r[6])
        if (num, pub) in seen:
            halt(f"duplicate enforcement notice key {(num, pub)}")
        seen.add((num, pub))
        out.append((num, pub, str(r[1]).strip(), text(r[4]), text(r[5]),
                    text(r[7]), text(r[8]), date(r[9]), text(r[2]), text(r[3]),
                    spec["edition_date"], spec["url"], spec["filename"],
                    spec["release_page_url"]))
    return out


def log_run(cur, n, reg, jud):
    cur.execute("""
        INSERT INTO pipeline_run_log
            (run_id, source_number, source_code, agent_name, status,
             rows_written, notes, started_at, completed_at)
        VALUES (gen_random_uuid(), %s, %s, %s, 'success', %s, %s, now(), now())
    """, ("24", SOURCE_CODE, "Source 24 - RSH register and judgements", n,
          f"Register snapshot {reg['snapshot_date']}, judgements edition "
          f"{jud['edition_date']}. Entity-level, no LA geography; deliberately "
          f"not wired into staging_la_signals."))


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--discover", action="store_true")
    ap.add_argument("--load", action="store_true")
    args = ap.parse_args()
    if not (args.discover or args.load):
        ap.error("choose --discover or --load")

    reg_spec = resolve_register()
    jud_spec = resolve_judgements()
    reg_rows = build_register(
        reg_spec, sheet_rows(fetch(reg_spec), REGISTER_HEADERS, "Registered Providers"))
    jud_path = fetch(jud_spec)
    jud_rows = build_judgements(
        jud_spec, sheet_rows(jud_path, JUDGEMENT_HEADERS, "Regulatory Judgements"))
    not_rows = build_notices(
        jud_spec, sheet_rows(jud_path, NOTICE_HEADERS, "Enforcement Notices"))

    print(f"register snapshot   : {reg_spec['snapshot_date']}  "
          f"({reg_spec['filename']})")
    print(f"  providers         : {len(reg_rows)}")
    import collections
    print(f"  designation       : "
          f"{dict(collections.Counter(r[4] for r in reg_rows))}")
    print(f"judgements edition  : {jud_spec['edition_date']}  "
          f"({jud_spec['filename']})")
    print(f"  judgements        : {len(jud_rows)}")
    print(f"  enforcement notices: {len(not_rows)}")
    graded = sum(1 for r in jud_rows if r[8])
    print(f"  with a governance grade: {graded}")

    conn = get_conn()
    cur = conn.cursor()
    try:
        if args.load:
            cur.execute(DDL)
            cur.execute(DDL_GUARDS)
            psycopg2.extras.execute_values(cur, REGISTER_UPSERT, reg_rows,
                                           page_size=500)
            psycopg2.extras.execute_values(cur, JUDGEMENT_UPSERT, jud_rows,
                                           page_size=500)
            psycopg2.extras.execute_values(cur, NOTICE_UPSERT, not_rows,
                                           page_size=500)
            log_run(cur, len(reg_rows) + len(jud_rows) + len(not_rows),
                    reg_spec, jud_spec)
            for t in (REGISTER_TABLE, JUDGEMENT_TABLE, NOTICE_TABLE):
                cur.execute(f"SELECT COUNT(*) FROM {t}")
                print(f"{t}: {cur.fetchone()[0]} rows")
            conn.commit()
            print("COMMITTED")
        else:
            conn.rollback()
            print("\nNothing committed.")
        return 0
    except Exception:
        conn.rollback()
        raise
    finally:
        cur.close()
        conn.close()


if __name__ == "__main__":
    sys.exit(main())
